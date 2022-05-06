
from cgi import print_form
from email.header import Header
import openpyxl
import pandas as pd
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter



teamsExcel = 'C:/Users/e.sarmiento/Desktop/Excel documents/team-members/team-members.xlsx'
masterExcel = 'C:/Users/e.sarmiento/Desktop/Excel documents/master05.xlsx'
agentsExcel =  'C:/Users/e.sarmiento/Desktop/Excel documents/agents/agents.xlsx'


#se lee team members y masterExcel
readTeamsExcel = pd.read_excel(teamsExcel)
readMasterExcelBdLogin = pd.read_excel(masterExcel, sheet_name='BD-Login')
readMasterExcelClouldtalk = pd.read_excel(masterExcel, sheet_name='Cloudtalk')

readagentsExcel = pd.read_excel(agentsExcel)
#se almacena teams
dataT =pd.DataFrame(readTeamsExcel[['Date','User ID','Name','Email']])
datafilteredteams = dataT.dropna(subset=['Date'])
print('1RA PARTE')
print(datafilteredteams)
dataT2 =pd.DataFrame(readTeamsExcel[['Group','Absence','Productive time','Unproductive time','Neutral time','Total DeskTime','Offline time','Private time','Arrived','Left','Late','Total time at work','Idle time','Extra hours before work','Extra hours after work','Hourly rate']])
datafilteredteams2 = dataT2.dropna(subset=['Group'])
print ('2da parte')
print(datafilteredteams2)
dataAgent =pd.DataFrame(readagentsExcel[['Date'	,'Agent','SIP Login time (sec)','Idle time (sec)','Ringing time (sec)','Talking time (sec)','Wrap up time (sec)','Inbound Calls','Outbound Calls']])
datafilteredAgent = dataAgent.dropna(subset=['Date'])
print(dataAgent)
#se lee master para obtener fila de bd-login
dataM = pd.DataFrame(readMasterExcelBdLogin)
datafiltered = dataM.dropna(subset=['Date'])
lastindex =datafiltered.index[-1]
#se lee master para obtener fila de cloudtalk
lastindexFound = pd.DataFrame(readMasterExcelClouldtalk)

lastindexCloud =lastindexFound.index[-1]
 #se pega el team member en master
#dataT.to_excel(masterExcel,startrow=lastindex+2,index=False,header=False,sheet_name='BD-Login')
with pd.ExcelWriter (masterExcel,mode="a",
    engine="openpyxl",
    if_sheet_exists="overlay",date_format="DD-MM-YYYY",) as writer:
   datafilteredteams.to_excel(writer,startrow=lastindex+2,index=False,header=False,sheet_name='BD-Login')
   datafilteredteams2.to_excel(writer,startrow=lastindex+2,index=False,header=False,sheet_name='BD-Login',startcol=5)
   dataAgent.to_excel(writer,index=False,startrow=lastindexCloud+2,header=False,sheet_name='Cloudtalk')
    






 




#archivo_excel_master =pd.read_excel('C:/Users/e.sarmiento/Desktop/Excel documents/master.xlsx')


#wb = load_workbook('C:/Users/e.sarmiento/Desktop/Excel documents/master.xlsx')
#ws = wb.active
#print(wb)

#wb2 = load_workbook('C:/Users/e.sarmiento/Desktop/Excel documents/team-members/team-members-export-20220505142801_2022-05-04_2022-05-04.xlsx')
#ws2 = wb2['Sheet1']

#for fila in range (2, ws2.max_row+1):
 # for colunma in range (1,5):
  #    char = get_column_letter(colunma)
   #   data = (ws2[char+str (fila)].value)
    #  print(data)

